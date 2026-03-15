from io import BytesIO
from pathlib import Path
from tempfile import gettempdir
from typing import Dict, List, Optional
from uuid import uuid4

import pandas as pd
from fastapi import Body, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from src.tally_tool.xml_importer import parse_tally_xml_content

app = FastAPI(title="Tally to Excel Backend", version="1.0.0")

# Development ke liye open rakha hai.
# Production me apna exact GitHub Pages domain hi allow karna.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

COLUMN_ALIASES = {
    "date": ["date", "voucher date", "dt"],
    "voucher_no": ["voucher no", "voucher number", "vch no", "vch number", "voucher_no", "voucher no."],
    "voucher_type": ["voucher type", "vch type", "voucher_type", "voucher typename"],
    "party_ledger": ["party ledger", "party", "party name", "party_ledger", "party ledger name"],
    "ledger": ["ledger", "ledger name", "particulars", "account", "account head"],
    "narration": ["narration", "remarks", "description"],
    "amount": ["amount", "amt", "value"],
    "debit": ["debit", "dr", "debit amount"],
    "credit": ["credit", "cr", "credit amount"],
    "gstin": ["gstin", "gst no", "gst"],
    "item_name": ["item name", "stock item", "item"],
    "quantity": ["qty", "quantity"],
    "rate": ["rate"],
}

REQUIRED_COLUMNS = [
    "date",
    "voucher_no",
    "voucher_type",
    "party_ledger",
    "ledger",
    "narration",
    "amount",
]

MEMORY_STORE: Dict[str, List[dict]] = {}
TEMP_OUTPUT_STORE: Dict[str, Path] = {}
OUTPUT_DIR = Path(gettempdir()) / "tally_to_excel_outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


class ProcessXMLRequest(BaseModel):
    file_id: str
    from_date: Optional[str] = None
    to_date: Optional[str] = None
    voucher_type: Optional[str] = None


def clean_col_name(col: str) -> str:
    return str(col).strip().lower().replace("_", " ")


def to_number(value):
    if pd.isna(value):
        return 0.0
    text = str(value).strip().replace(",", "")
    if text == "":
        return 0.0

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    try:
        num = float(text)
        return -num if negative else num
    except Exception:
        return 0.0


def parse_date_series(series: pd.Series) -> pd.Series:
    # Pehle normal parse
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)

    # Jo parse nahi hua aur 8 digit format hai, use YYYYMMDD maan lo
    mask = parsed.isna()
    if mask.any():
        raw = series.astype(str).str.strip()
        eight_digit = raw.str.match(r"^\d{8}$", na=False)
        idx = mask & eight_digit
        if idx.any():
            parsed.loc[idx] = pd.to_datetime(raw.loc[idx], format="%Y%m%d", errors="coerce")

    return parsed


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    original_cols = list(df.columns)
    cleaned_map = {col: clean_col_name(col) for col in original_cols}

    rename_map = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for original, cleaned in cleaned_map.items():
            if cleaned in aliases:
                rename_map[original] = canonical
                break

    df = df.rename(columns=rename_map)

    # Ensure columns exist
    for col in set(REQUIRED_COLUMNS + ["debit", "credit", "gstin", "item_name", "quantity", "rate"]):
        if col not in df.columns:
            df[col] = None

    # Dates
    df["date"] = parse_date_series(df["date"])

    # Numbers
    if "debit" in df.columns:
        df["debit"] = df["debit"].apply(to_number)
    if "credit" in df.columns:
        df["credit"] = df["credit"].apply(to_number)

    if "amount" in df.columns:
        df["amount"] = df["amount"].apply(to_number)

    # Agar amount khaali ho to debit/credit se derive karo
    if df["amount"].fillna(0).abs().sum() == 0:
        if "debit" in df.columns and "credit" in df.columns:
            df["amount"] = df["debit"].fillna(0) - df["credit"].fillna(0)
        elif "debit" in df.columns:
            df["amount"] = df["debit"].fillna(0)
        elif "credit" in df.columns:
            df["amount"] = -df["credit"].fillna(0)

    # Text cleanup
    for col in ["voucher_no", "voucher_type", "party_ledger", "ledger", "narration", "gstin", "item_name"]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    if "quantity" in df.columns:
        df["quantity"] = df["quantity"].apply(to_number)
    if "rate" in df.columns:
        df["rate"] = df["rate"].apply(to_number)

    # Display amount
    df["abs_amount"] = df["amount"].abs()

    return df


async def load_file_to_df(file: UploadFile) -> pd.DataFrame:
    filename = file.filename.lower()
    content = await file.read()

    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(content))
            return normalize_columns(df)

        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            df = pd.read_excel(BytesIO(content))
            return normalize_columns(df)

        if filename.endswith(".xls"):
            df = pd.read_excel(BytesIO(content), engine="xlrd")
            return normalize_columns(df)

        if filename.endswith(".xml"):
            xml_text = content.decode("utf-8", errors="ignore")
            return normalize_columns(parse_tally_xml_content(xml_text))

        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Use CSV, XLSX, XLSM, XLS, or XML.",
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File reading failed: {str(e)}")


def apply_filters(
    df: pd.DataFrame,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    voucher_types: Optional[str] = None,
) -> pd.DataFrame:
    result = df.copy()

    if from_date:
        fd = pd.to_datetime(from_date, errors="coerce")
        if pd.notna(fd):
            result = result[result["date"] >= fd]

    if to_date:
        td = pd.to_datetime(to_date, errors="coerce")
        if pd.notna(td):
            result = result[result["date"] <= td]

    if voucher_types:
        vt_list = [x.strip().lower() for x in voucher_types.split(",") if x.strip()]
        if vt_list:
            result = result[
                result["voucher_type"].str.lower().apply(lambda x: any(v in x for v in vt_list))
            ]

    return result


def dataframe_preview_payload(df: pd.DataFrame) -> dict:
    preview_df = df.copy()
    if "date" in preview_df.columns:
        preview_df["date"] = preview_df["date"].dt.strftime("%Y-%m-%d")

    preview_df = preview_df.where(pd.notna(preview_df), None)
    preview_df = preview_df.head(100)

    return {
        "preview": preview_df.to_dict(orient="records"),
        "columns": list(preview_df.columns),
        "total_rows": int(len(df)),
    }


def save_excel_to_temp(excel_file: BytesIO) -> str:
    output_file_id = str(uuid4())
    output_path = OUTPUT_DIR / f"{output_file_id}.xlsx"
    with output_path.open("wb") as f:
        f.write(excel_file.getvalue())

    TEMP_OUTPUT_STORE[output_file_id] = output_path
    return output_file_id

def build_purchase_register(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["voucher_type"].str.lower().str.contains("purchase", na=False)
    cols = ["date", "voucher_no", "voucher_type", "party_ledger", "ledger", "narration", "abs_amount"]
    out = df.loc[mask, cols].copy()
    out = out.rename(columns={"abs_amount": "amount"})
    return out.sort_values(["date", "voucher_no"], na_position="last")


def build_sales_register(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["voucher_type"].str.lower().str.contains("sales|sale", na=False)
    cols = ["date", "voucher_no", "voucher_type", "party_ledger", "ledger", "narration", "abs_amount"]
    out = df.loc[mask, cols].copy()
    out = out.rename(columns={"abs_amount": "amount"})
    return out.sort_values(["date", "voucher_no"], na_position="last")


def build_journal_register(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["voucher_type"].str.lower().str.contains("journal", na=False)
    cols = ["date", "voucher_no", "voucher_type", "party_ledger", "ledger", "narration", "amount", "abs_amount"]
    out = df.loc[mask, cols].copy()
    out = out.sort_values(["date", "voucher_no"], na_position="last")
    return out


def build_ledger_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["ledger", "entry_count", "total_amount"])

    out = (
        df.groupby("ledger", dropna=False)
        .agg(entry_count=("ledger", "count"), total_amount=("abs_amount", "sum"))
        .reset_index()
        .sort_values("total_amount", ascending=False)
    )
    return out


def build_gst_summary(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    work["ledger_lower"] = work["ledger"].str.lower()

    def total_for(keyword_list: List[str]) -> float:
        mask = work["ledger_lower"].apply(lambda x: any(k in x for k in keyword_list))
        return float(work.loc[mask, "abs_amount"].sum())

    summary = pd.DataFrame(
        [
            {"particulars": "CGST", "amount": total_for(["cgst"])},
            {"particulars": "SGST", "amount": total_for(["sgst"])},
            {"particulars": "IGST", "amount": total_for(["igst"])},
            {"particulars": "CESS", "amount": total_for(["cess"])},
        ]
    )

    summary.loc[len(summary)] = {"particulars": "Total GST", "amount": float(summary["amount"].sum())}

    return summary


def format_workbook(writer, df_map: dict):
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")

    for sheet_name in df_map:
        ws = writer.sheets[sheet_name]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Auto width
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)


def build_excel_file(df: pd.DataFrame, report_type: str) -> BytesIO:
    output = BytesIO()

    master_cols = ["date", "voucher_no", "voucher_type", "party_ledger", "ledger", "narration", "amount", "abs_amount"]

    purchase_df = build_purchase_register(df)
    sales_df = build_sales_register(df)
    journal_df = build_journal_register(df)
    ledger_summary_df = build_ledger_summary(df)
    gst_summary_df = build_gst_summary(df)

    report_type = (report_type or "all").strip().lower()

    sheet_map = {}

    if report_type == "purchase_register":
        sheet_map["Purchase Register"] = purchase_df
    elif report_type == "sales_register":
        sheet_map["Sales Register"] = sales_df
    elif report_type == "journal_register":
        sheet_map["Journal Register"] = journal_df
    elif report_type == "ledger_summary":
        sheet_map["Ledger Summary"] = ledger_summary_df
    elif report_type == "gst_summary":
        sheet_map["GST Summary"] = gst_summary_df
    else:
        sheet_map["Master Data"] = df[master_cols].sort_values(["date", "voucher_no"], na_position="last")
        sheet_map["Purchase Register"] = purchase_df
        sheet_map["Sales Register"] = sales_df
        sheet_map["Journal Register"] = journal_df
        sheet_map["Ledger Summary"] = ledger_summary_df
        sheet_map["GST Summary"] = gst_summary_df

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheet_map.items():
            sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
        format_workbook(writer, sheet_map)

    output.seek(0)
    return output


@app.get("/health")
def health():
    return {"status": "ok", "message": "Backend running"}


@app.post("/preview")
async def preview_file(
    file: UploadFile = File(...),
    from_date: Optional[str] = Form(None),
    to_date: Optional[str] = Form(None),
    voucher_types: Optional[str] = Form(None),
):
    df = await load_file_to_df(file)
    df = apply_filters(df, from_date=from_date, to_date=to_date, voucher_types=voucher_types)
    payload = dataframe_preview_payload(df)

    return JSONResponse(
        {
            "rows": payload["preview"],
            "columns": payload["columns"],
            "total_rows": payload["total_rows"],
        }
    )


@app.post("/upload-xml-preview")
async def upload_xml_preview(file: UploadFile = File(...)):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xml"):
        raise HTTPException(status_code=400, detail="Only XML file is supported on this endpoint.")

    try:
        content = await file.read()
        xml_text = content.decode("utf-8", errors="ignore")
        df = parse_tally_xml_content(xml_text)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    input_file_id = str(uuid4())
    MEMORY_STORE[input_file_id] = df.to_dict(orient="records")
    payload = {
        "preview": df.head(100).fillna("").to_dict(orient="records"),
        "columns": list(df.columns),
        "total_rows": int(len(df)),
    }
    return {
        "file_id": input_file_id,
        **payload,
    }


@app.post("/process")
async def process_file(
    file: UploadFile = File(...),
    report_type: str = Form("all"),
    from_date: Optional[str] = Form(None),
    to_date: Optional[str] = Form(None),
    voucher_types: Optional[str] = Form(None),
):
    df = await load_file_to_df(file)
    df = apply_filters(df, from_date=from_date, to_date=to_date, voucher_types=voucher_types)

    if df.empty:
        raise HTTPException(status_code=400, detail="No data found after applying filters.")

    excel_file = build_excel_file(df, report_type=report_type)

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="tally_report_{report_type}.xlsx"'},
    )


@app.post("/process-xml")
async def process_xml(request: ProcessXMLRequest = Body(...)):
    if request.file_id not in MEMORY_STORE:
        raise HTTPException(status_code=404, detail="Uploaded file not found. Please preview XML first.")

    filtered_df = pd.DataFrame(MEMORY_STORE[request.file_id])

    if request.from_date:
        from_date = pd.to_datetime(request.from_date).date()
        filtered_df = filtered_df[
            filtered_df["Date"].apply(lambda x: bool(x) and pd.to_datetime(x).date() >= from_date)
        ]

    if request.to_date:
        to_date = pd.to_datetime(request.to_date).date()
        filtered_df = filtered_df[
            filtered_df["Date"].apply(lambda x: bool(x) and pd.to_datetime(x).date() <= to_date)
        ]

    if request.voucher_type and request.voucher_type.lower() != "all":
        filtered_df = filtered_df[
            filtered_df["Voucher Type"].astype(str).str.lower() == request.voucher_type.lower()
        ]

    output_file_id = str(uuid4())
    output_path = OUTPUT_DIR / f"{output_file_id}.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        filtered_df.to_excel(writer, index=False, sheet_name="Tally Data")

    TEMP_OUTPUT_STORE[output_file_id] = output_path

    return {
        "status": "success",
        "file_id": output_file_id,
        "rows_after_filter": int(len(filtered_df)),
    }


@app.get("/download/{file_id}")
async def download_file(file_id: str):
    file_path = TEMP_OUTPUT_STORE.get(file_id)
    if not file_path or not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found or expired.")

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"tally_report_{file_id}.xlsx",
    )
