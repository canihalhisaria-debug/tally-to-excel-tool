"""Excel export utilities using openpyxl formatting primitives."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import ReportConfig

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)


def _write_dataframe_sheet(workbook: Workbook, name: str, data: pd.DataFrame, config: ReportConfig) -> None:
    ws = workbook.create_sheet(title=name)

    if data.empty:
        ws.append(["No records found"])
        return

    ws.append(list(data.columns))
    for row in data.itertuples(index=False):
        ws.append(list(row))

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, pd.Timestamp):
                cell.number_format = config.date_output_format
            elif isinstance(cell.value, (int, float)):
                header = str(ws.cell(1, cell.column).value or "").lower()
                if any(token in header for token in ("debit", "credit", "amount", "gst", "expense", "total")):
                    cell.number_format = config.currency_format

    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def build_excel_report(reports: dict[str, pd.DataFrame], config: ReportConfig) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    ordered_names = [name for name in config.default_sheet_order if name in reports]
    ordered_names.extend([name for name in reports if name not in ordered_names])

    for report_name in ordered_names:
        _write_dataframe_sheet(workbook, report_name, reports[report_name], config)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
